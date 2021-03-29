from openpyxl import Workbook
import time
import os
def main(number,time):
    # 創建一個空白活頁簿物件
    wb = Workbook()
    # 選取正在工作中的表單
    ws = wb.active
    # 指定值給 A1 儲存格
    ws['A1'] = '學號'
    ws['B1'] = '簽到時間'
    # 向下新增一列並連續插入值
    ws.append([number, time])
    # 儲存成 create_sample.xlsx 檔案
    wb.save('./excel_test/create_sample.xlsx')

# 取得 struct_time 格式的時間
t = time.localtime()
# 依指定格式輸出
Now_time = time.strftime("%H:%M:%S", t)
main("M10907324",Now_time)

